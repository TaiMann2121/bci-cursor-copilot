"""
profile_sources.py
==================
Quantify the trajectory PROFILE of each data source (real EEGK, calibrated sim,
surrogate if present) so we can specify what "matching the real data" means.

Produces:
  results/source_profiles.xlsx  — metadata + per-subject + per-direction tables
  (a companion spec doc is generated separately)

Motivation (supervisor, 7/13): the prerequisite for testing "does added data
help" is that the added data actually matches the real profile. That requires a
written-down core parameter set and the per-subject / per-direction distributions
to compare against. This script IS that measurement, and is regenerable when new
real EEGK data lands.

The core parameter set (8 params, 3 families):
  SPATIAL   : endpoint_radius, endpoint_angle_error
  TEMPORAL  : length_ticks, dwell_ticks
  KINEMATIC : step_mag_mean, step_mag_std, wander_index, reversal_rate

Each is reported per subject and per direction, with real as the reference and
each synthetic source shown beside it. A real-vs-sim comparison flags which
properties are decoder-shared (real≈sim) vs subject-specific (real≠sim), since
sim re-decodes the same trials through the same decoder.
"""
from __future__ import annotations
import argparse, warnings
from collections import defaultdict
from pathlib import Path
import numpy as np

import copilot_dataset as cd
import sim_scaling as ss

warnings.filterwarnings("ignore")

DWELL_RADIUS = 0.05
PARAMS = ["endpoint_radius", "endpoint_angle_error_deg", "length_ticks", "dwell_ticks",
          "step_mag_mean", "step_mag_std", "wander_index", "reversal_rate"]
FAMILY = {"endpoint_radius":"spatial","endpoint_angle_error_deg":"spatial",
          "length_ticks":"temporal","dwell_ticks":"temporal",
          "step_mag_mean":"kinematic","step_mag_std":"kinematic",
          "wander_index":"kinematic","reversal_rate":"kinematic"}
SUBJECTS = None  # derived from real data at load time (see load_sources); excludes sim-only subjects automatically
DIRS = ["NW","N","NE","W","E","SW","S","SE"]  # display order


def trajectory_params(t) -> dict:
    p = t.pos
    steps = p[1:] - p[:-1]
    slen = np.linalg.norm(steps, axis=1)
    r = np.linalg.norm(p, axis=1)
    k = 0
    while k < len(r) and r[k] < DWELL_RADIUS:
        k += 1
    rev = ns = 0
    for i in range(len(steps) - 1):
        a, b = steps[i], steps[i+1]
        na, nb = np.linalg.norm(a), np.linalg.norm(b)
        if na > 1e-9 and nb > 1e-9:
            if np.dot(a, b) / (na * nb) < 0:
                rev += 1
            ns += 1
    endr = float(np.linalg.norm(p[-1]))
    tgt = cd.TARGET_POS[t.target_label]
    ta = np.arctan2(tgt[1], tgt[0]); ea = np.arctan2(p[-1][1], p[-1][0])
    aerr = float(abs(np.degrees(np.arctan2(np.sin(ea - ta), np.cos(ea - ta)))))
    return {
        "endpoint_radius": endr,
        "endpoint_angle_error_deg": aerr,
        "length_ticks": float(t.n_ticks),
        "dwell_ticks": float(k),
        "step_mag_mean": float(slen.mean()) if len(slen) else 0.0,
        "step_mag_std": float(slen.std()) if len(slen) else 0.0,
        "wander_index": float(slen.sum() / max(endr, 1e-6)),
        "reversal_rate": float(rev / max(ns, 1)),
    }


def summarize(trajs) -> dict:
    """mean of each param over a set of trajectories (median for wander, which is skewed)."""
    if not trajs:
        return {p: np.nan for p in PARAMS}
    rows = [trajectory_params(t) for t in trajs]
    out = {}
    for p in PARAMS:
        vals = np.array([r[p] for r in rows])
        out[p] = float(np.median(vals)) if p == "wander_index" else float(vals.mean())
    return out


def load_sources(repo_root, seed, match, add_dwell):
    global SUBJECTS
    real = cd.load_source("eegk_real", repo_root=repo_root)
    SUBJECTS = sorted({t.subject_id for t in real})   # derive from real; sim-only subjects excluded downstream
    print(f"subjects (from real data): {SUBJECTS}")
    sim_raw = cd.load_source("eegk_sim", repo_root=repo_root)
    sim = ss.scale_sim_to_real(sim_raw, real, match=match)
    if add_dwell:
        sim = ss.add_dwell_to_sim(sim, real, seed=seed)
    sim = [t for t in sim if t.subject_id in SUBJECTS]
    sources = {"eegk_real": real, "eegk_sim_calibrated": sim}
    sp = Path(repo_root) / "data/surrogate/surrogate_trajectories.csv"
    if sp.exists():
        surr = [t for t in cd.load_csv_file(str(sp)) if t.subject_id in SUBJECTS]
        if surr:
            sources["eegk_surrogate"] = surr
    return sources


def build_workbook(sources, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    HDR_FILL = PatternFill("solid", fgColor="1F5C8B")
    SUB_FILL = PatternFill("solid", fgColor="DCE6F1")
    ALT_FILL = PatternFill("solid", fgColor="F2F5F8")
    REF_FILL = PatternFill("solid", fgColor="FFF4E0")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    bold = Font(name=FONT, bold=True, size=10)
    reg = Font(name=FONT, size=10)
    mute = Font(name=FONT, size=9, italic=True, color="666666")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")

    wb = openpyxl.Workbook()

    def style_header(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL; cell.font = hdr_font; cell.alignment = ctr; cell.border = border

    # ---------- Sheet 1: README ----------
    ws = wb.active; ws.title = "README"
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 90
    ws["A1"] = "Trajectory Source Profiles"; ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F5C8B")
    lines = [
        ("Purpose", "Quantify each data source's trajectory profile so 'matching the real data' is measurable. Real EEGK is the reference; synthetic sources are shown beside it."),
        ("Provisional", "Numbers reflect the current 5 subjects (S01,S02,S04,S05,S07). Re-run this script when more real EEGK data lands."),
        ("Tabs", "metadata_counts = trial counts. per_subject / per_direction = the 8 profile parameters. match_subject = go/no-go for adding sim to a subject (PRIMARY). match_direction = which directions sim matches/misses (diagnostic)."),
        ("match metric", "|sim_mean - real_mean| / real_std, computed WITHIN each subject/direction on its own trials. <1 = sim's mean sits inside real's own trial-to-trial spread (tight match); >2 = outside it (mismatch). Purple '>=10' = real had ~no spread on that param (e.g. a subject whose trajectories are all the same length), so there is no within-group scale to judge against. This is the per-subject decision metric: does sim look like it came from THIS subject's data. Heuristic, not a significance test."),
        ("Core parameters", "SPATIAL: endpoint_radius, endpoint_angle_error_deg | TEMPORAL: length_ticks, dwell_ticks | KINEMATIC: step_mag_mean, step_mag_std, wander_index, reversal_rate"),
        ("wander_index", "path length / straight-line distance; reported as MEDIAN (skewed by a chaotic tail). All other params are means."),
        ("Sources", ", ".join(sources.keys())),
    ]
    r = 3
    for k, v in lines:
        ws.cell(row=r, column=1, value=k).font = bold
        c = ws.cell(row=r, column=2, value=v); c.font = reg; c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    # ---------- Sheet 2: metadata_counts ----------
    ws = wb.create_sheet("metadata_counts")
    ws.cell(row=1, column=1, value="Trial counts per subject x direction (source: eegk_real)").font = bold
    hdr_row = 3
    ws.cell(row=hdr_row, column=1, value="Subject")
    for j, d in enumerate(DIRS):
        ws.cell(row=hdr_row, column=2+j, value=d)
    ws.cell(row=hdr_row, column=2+len(DIRS), value="TOTAL")
    style_header(ws, hdr_row, 2+len(DIRS))
    real = sources["eegk_real"]
    name_to_lbl = {v: k for k, v in cd.DIR_NAMES.items()}
    counts = defaultdict(lambda: defaultdict(int))
    for t in real:
        counts[t.subject_id][cd.DIR_NAMES[t.target_label]] += 1
    rr = hdr_row + 1
    for si, s in enumerate(SUBJECTS):
        ws.cell(row=rr, column=1, value=s).font = bold
        tot = 0
        for j, d in enumerate(DIRS):
            v = counts[s][d]; tot += v
            cell = ws.cell(row=rr, column=2+j, value=v); cell.font = reg; cell.alignment = ctr; cell.border = border
        tc = ws.cell(row=rr, column=2+len(DIRS), value=tot); tc.font = bold; tc.alignment = ctr; tc.border = border
        if si % 2 == 1:
            for c in range(1, 3+len(DIRS)):
                if ws.cell(row=rr, column=c).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=rr, column=c).fill = ALT_FILL
        ws.cell(row=rr, column=1).border = border
        rr += 1
    # direction totals
    ws.cell(row=rr, column=1, value="TOTAL").font = bold
    grand = 0
    for j, d in enumerate(DIRS):
        col_tot = sum(counts[s][d] for s in SUBJECTS); grand += col_tot
        cell = ws.cell(row=rr, column=2+j, value=col_tot); cell.font = bold; cell.alignment = ctr; cell.border = border; cell.fill = SUB_FILL
    gc = ws.cell(row=rr, column=2+len(DIRS), value=grand); gc.font = bold; gc.alignment = ctr; gc.border = border; gc.fill = SUB_FILL
    ws.cell(row=rr, column=1).fill = SUB_FILL; ws.cell(row=rr, column=1).border = border
    ws.cell(row=rr+2, column=1, value="Note: real EEGK is direction-imbalanced; imbalance is itself a property added data should be aware of.").font = mute
    ws.column_dimensions["A"].width = 10
    for j in range(len(DIRS)+1):
        ws.column_dimensions[get_column_letter(2+j)].width = 8

    # ---------- helper to write a param table (rows = subject/dir x source) ----------
    def param_sheet(sheet_name, group_key_fn, group_labels, title):
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=title).font = bold
        hdr = 3
        ws.cell(row=hdr, column=1, value="Group")
        ws.cell(row=hdr, column=2, value="Source")
        for j, p in enumerate(PARAMS):
            ws.cell(row=hdr, column=3+j, value=p)
        style_header(ws, hdr, 2+len(PARAMS))
        rr = hdr + 1
        for gi, g in enumerate(group_labels):
            for src_name, trajs in sources.items():
                sub = [t for t in trajs if group_key_fn(t) == g]
                summ = summarize(sub)
                is_real = src_name == "eegk_real"
                gc = ws.cell(row=rr, column=1, value=g if src_name == list(sources)[0] else "")
                gc.font = bold; gc.alignment = lft; gc.border = border
                sc_ = ws.cell(row=rr, column=2, value=src_name); sc_.font = (bold if is_real else reg); sc_.border = border; sc_.alignment = lft
                for j, p in enumerate(PARAMS):
                    v = summ[p]
                    cell = ws.cell(row=rr, column=3+j, value=round(v, 3) if not np.isnan(v) else None)
                    cell.font = reg; cell.alignment = ctr; cell.border = border
                    if is_real:
                        cell.fill = REF_FILL
                if is_real:
                    ws.cell(row=rr,column=1).fill = REF_FILL; ws.cell(row=rr,column=2).fill = REF_FILL
                rr += 1
            rr += 0
        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 20
        for j in range(len(PARAMS)):
            ws.column_dimensions[get_column_letter(3+j)].width = 15
        ws.freeze_panes = "C4"
        return ws

    param_sheet("per_subject", lambda t: t.subject_id, SUBJECTS,
                "Profile parameters per subject (real = highlighted reference)")
    param_sheet("per_direction", lambda t: cd.DIR_NAMES[t.target_label], DIRS,
                "Profile parameters per direction, pooled over subjects (real = highlighted reference)")

    real = sources["eegk_real"]; sim = sources.get("eegk_sim_calibrated", [])

    # ---------- within-group match: does sim match THIS group's own real data? ----------
    # PRIMARY metric for a per-subject task. For each group (subject or direction):
    #   |sim_mean - real_mean| / real_std_within_that_group
    # i.e. how far sim's average sits from real's average, measured in units of that
    # group's OWN trial-to-trial spread. <1 = sim's mean is inside real's own
    # distribution (tight match); >2 = outside where real's data usually falls.
    # Note the opposite failure mode from the between-group metric: a group whose real
    # data is very CONSISTENT (tiny std) makes even small sim gaps look large; where
    # real_std ~ 0 the ratio is undefined and shown as ">=CAP" / n-a.
    STD_FLOOR = 1e-6
    CAP = 10.0

    def per_trial_vals(trajs, param):
        return np.array([trajectory_params(t)[param] for t in trajs])

    def within_group_sheet(sheet_name, groups, key_fn, group_label, title):
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=title).font = bold
        purpose = ("ANSWERS: should I add sim to a given subject's training data? (go/no-go)"
                   if group_label == "subject"
                   else "ANSWERS: for which target directions does sim match / miss real? (diagnostic)")
        ws.cell(row=2, column=1, value=purpose).font = Font(name=FONT, bold=True, size=10, color="1F5C8B")
        ws.cell(row=3, column=1, value=("Ratio = |sim_mean - real_mean| / real_std, computed WITHIN each "
                f"{group_label} on its own trials. <1 = sim's mean is inside real's own distribution "
                "(tight match); >2 = outside it (mismatch for this "+group_label+").")).font = mute
        ws.cell(row=4, column=1, value=("Consistency caution: a "+group_label+" whose real data is very "
                "uniform (near-zero std) inflates the ratio for tiny gaps; '>=10' or 'n/a' means real "
                "had ~no spread on that param, so there is no within-group scale to judge against.")).font = mute
        hdr = 6
        ws.cell(row=hdr, column=1, value=group_label.capitalize())
        for j, p in enumerate(PARAMS):
            ws.cell(row=hdr, column=2+j, value=p)
        style_header(ws, hdr, 1+len(PARAMS))
        rr = hdr + 1
        for g in groups:
            rtr = [t for t in real if key_fn(t) == g]
            str_ = [t for t in sim if key_fn(t) == g]
            ws.cell(row=rr, column=1, value=g).font = bold
            ws.cell(row=rr, column=1).border = border; ws.cell(row=rr, column=1).alignment = ctr
            for j, p in enumerate(PARAMS):
                if not rtr or not str_:
                    val, ratio = None, np.nan
                else:
                    rv = per_trial_vals(rtr, p); sv = per_trial_vals(str_, p)
                    rstd = rv.std()
                    if rstd < STD_FLOOR:
                        ratio = np.inf
                        val = ">=%g" % CAP
                    else:
                        ratio = abs(sv.mean() - rv.mean()) / rstd
                        val = round(min(ratio, CAP), 2) if ratio < CAP else ">=%g" % CAP
                cell = ws.cell(row=rr, column=2+j, value=val)
                cell.font = reg; cell.alignment = ctr; cell.border = border
                if np.isnan(ratio):
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")
                elif np.isinf(ratio):
                    cell.fill = PatternFill("solid", fgColor="E4D6EA")   # undefined (no real spread)
                elif ratio < 1.0:
                    cell.fill = PatternFill("solid", fgColor="D6EBD6")
                elif ratio > 2.0:
                    cell.fill = PatternFill("solid", fgColor="F5D6D6")
                else:
                    cell.fill = PatternFill("solid", fgColor="FBF0D6")
            rr += 1
        ws.cell(row=rr+1, column=1,
                value="green <1 (sim mean inside real's own spread)   yellow 1-2   red >2 (mismatch)   "
                      "purple = no real spread to judge   grey = n/a").font = mute
        ws.column_dimensions["A"].width = 12
        for j in range(len(PARAMS)):
            ws.column_dimensions[get_column_letter(2+j)].width = 15
        ws.freeze_panes = "B7"

    within_group_sheet("match_subject", SUBJECTS,
                       lambda t: t.subject_id, "subject",
                       "Sim match to each subject's OWN real data (primary per-subject metric)")
    within_group_sheet("match_direction", DIRS,
                       lambda t: cd.DIR_NAMES[t.target_label], "direction",
                       "Sim match to each direction's OWN real data (pooled over subjects)")

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo_root", default=".")
    ap.add_argument("--seed", type=int, default=0)
    ap.add_argument("--match", default="radius", choices=["radius", "step"])
    ap.add_argument("--no_dwell", action="store_true")
    ap.add_argument("--out", default="results/source_profiles.xlsx")
    args = ap.parse_args()

    sources = load_sources(args.repo_root, args.seed, args.match, not args.no_dwell)
    print(f"sources: {list(sources.keys())}")
    for n, t in sources.items():
        print(f"  {n}: {len(t)} trajectories")
    out = Path(args.repo_root) / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(sources, str(out))
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
